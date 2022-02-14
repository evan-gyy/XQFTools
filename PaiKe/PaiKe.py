# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author: gyy
import openpyxl
import gc
import os
import datetime


class ClassPlan:
    def __init__(self):
        self.from_file_name = ""
        self.to_file_name = ""
        self.keyword_list = []
        self.plan_dict = {}
        self.stop_words = ['星', '一', '二', '三', '四', '五', '六', '日']

    def find_excel(self):
        excel_list = []
        for root, dirs, files in os.walk("."):
            for f in files:
                if '.xlsx' in f:
                    excel_list.append(f)
        excel_list = list(set(excel_list))
        print("检测到以下Excel文件：")
        for f in excel_list:
            print("{}：{}".format(excel_list.index(f), f))
        while True:
            try:
                self.from_file_name = excel_list[int((input("请输入排课源文件序号：")))]
                self.to_file_name = excel_list[int(input("请输入导出文件序号："))]
                break
            except:
                print("发生错误：请正确输入文件前的序号（0-n）")

    def validate(self, text):
        try:
            datetime.datetime.strptime(text, '%Y/%m/%d')
            return True
        except:
            return False

    def get_keyword(self, ws):
        max_row = ws.max_row
        for row in range(1, max_row):
            first_cell = ws.cell(row, 2).value
            if self.validate(first_cell):
                continue
            if first_cell and first_cell[0] in self.stop_words:
                continue
            for col in range(2, 8):
                cell = ws.cell(row, col).value
                if cell:
                    self.keyword_list.append(cell)
        self.keyword_list = list(set(self.keyword_list))

    def get_date(self, ws, key_word):
        max_row = ws.max_row
        # print(f"正在提取关键词：“{key_word}”")

        date_list = []
        time_list = []

        for row in range(1, max_row):
            for col in range(1, 8):
                cell = ws.cell(row, col).value
                if cell == key_word:
                    time = ws.cell(row, 1).value
                    time = time.replace("~", "-").replace(".", ":")
                    # print(time)
                    time_list.append(time)
                    for i in range(1, 10):
                        temp = ws.cell(row-i, col).value
                        if self.validate(temp):
                            # print(temp)
                            date_list.append(temp)
                            break
        # print("提取完成")
        return date_list, time_list

    def to_file(self):
        wb = openpyxl.load_workbook(self.to_file_name)
        for keyword in self.keyword_list:
            ws = wb.copy_worksheet(wb.worksheets[0])
            ws.title = keyword
            date_list = self.plan_dict[keyword]['date']
            time_list = self.plan_dict[keyword]['time']
            for i in range(len(date_list)):
                ws.cell(i+2, 1).value = i + 1
                ws.cell(i+2, 2).value = date_list[i]
                ws.cell(i+2, 3).value = time_list[i]
        for ws in wb.worksheets:
            if ws.title not in self.keyword_list:
                wb.remove(ws)
        wb.save(self.to_file_name)
        print(f"已写入文件：“{self.to_file_name}”")
        del wb, ws
        gc.collect()

    def run(self):
        self.find_excel()
        wb = openpyxl.load_workbook(self.from_file_name)
        ws = wb.active
        self.get_keyword(ws)
        for keyword in self.keyword_list:
            dl, tl = self.get_date(ws, keyword)
            self.plan_dict[keyword] = {
                'date': dl,
                'time': tl
            }
        # print(self.plan_dict)
        self.to_file()
        del wb, ws
        gc.collect()


def main():
    cp = ClassPlan()
    cp.run()
    input("运行完成")

main()
