# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author: gyy
import pandas as pd
import openpyxl
import os
import gc


class RenameNumber:
    def __init__(self):
        self.QUESTION = ("_word_question", "_T")
        self.ANSWER = ("_word_answer", "_A")
        self.RESOLVE = ("_word_resolve", "_R")
        self.doc_info = {}
        self.ql, self.al, self.rl = [], [], []
        self.content_cols = ['章', '节', '小节', '4级目录', '视频编号', '习题编号', '显示习题序号', '讲解内容', '排序', '目录关联知识点编号', '难度系数', 'ppt资源编号',
                        'pdf资源编号-学生', '空行数', 'pdf资源编号-老师', '讲义标题', 'pdf资源编号-出版', '目录ID', '生效时间', '关联习题集ID', '目录模式',
                        '视频数量(仅查看非导入属性)', '是否有效(仅查看非导入属性)', '英语文章ID(仅查看非导入属性)', 'word导入数量(仅查看非导入属性)', '年份(仅查看非导入属性)',
                        '来源(仅查看非导入属性)', '习题知识点编号(仅查看非导入属性)']

    def find_file(self, type):
        file_list = []
        for f in os.listdir("."):
            if type in f:
                file_list.append(f)
        print("检测到以下{}文件: ".format(type))
        for f in file_list:
            print("{}: {}".format(file_list.index(f), f))
        while True:
            try:
                target = file_list[int(input("请输入文件序号："))]
                return target
            except:
                print("发生错误：请正确输入文件前的序号（0-n）")

    def find_func(self, mode, ql, al, rl):
        """ functions for different mode """
        def case_func_1(ql, al, rl):
            result = []
            for i in range(len(ql)):
                result.append(ql[i])
                result.append(al[i])
                result.append(rl[i])
            return result

        def case_func_2(ql, al, rl):
            result = []
            for i in range(len(ql)):
                result.append(ql[i])
                result.append(rl[i])
                result.append(al[i])
            return result

        def case_func_3(ql, al, rl):
            result = []
            for i in range(len(ql)):
                result.append(ql[i])
            for i in range(len(al)):
                result.append(al[i])
                result.append(rl[i])
            return result

        def case_func_4(ql, al, rl):
            result = []
            result += ql
            for i in range(len(al)):
                result.append(rl[i])
                result.append(al[i])
            return result

        def case_func_5(ql, al, rl):
            result = []
            result += ql
            result += al
            result += rl
            return result

        def case_func_6(ql, al, rl):
            return ql

        def case_func_7(ql, al, rl):
            return al

        def case_func_8(ql, al, rl):
            return rl

        def case_func_9(ql, al, rl):
            result = []
            for i in range(len(ql)):
                result.append(ql[i])
                result.append(al[i])
            return result

        def case_func_10(ql, al, rl):
            result = []
            result += ql
            result += al
            return result

        func_map = {
            1: case_func_1,
            2: case_func_2,
            3: case_func_3,
            4: case_func_4,
            5: case_func_5,
            6: case_func_6,
            7: case_func_7,
            8: case_func_8,
            9: case_func_9,
            10: case_func_10,
        }
        if mode not in func_map:
            return 0

        return func_map[mode](ql, al, rl)

    def get_names(self):
        df = pd.read_excel(self.find_file('.xlsx'), usecols=[0])
        df.dropna(inplace=True)
        df['question'] = df.iloc[:, 0].apply(lambda x: x + self.QUESTION)
        df['answer'] = df.iloc[:, 0].apply(lambda x: x + self.ANSWER)
        df['resolve'] = df.iloc[:, 0].apply(lambda x: x + self.RESOLVE)
        ques_list = df['question'].tolist()
        ans_list = df['answer'].tolist()
        res_list = df['resolve'].tolist()
        return ques_list, ans_list, res_list

    def get_document_info(self):
        df = pd.read_excel(self.find_file('.xlsx'))
        for i, row in df.iterrows():
            if row.isnull().any():
                print('有效目录: ', i)
                break

            keyword = row['关键词']
            mode = int(row['模式'])
            volume = int(row['题量'])
            num_list = [f"{keyword}{i:0>3d}" for i in range(1, volume + 1)]
            final_list = self.run_mode(mode, [n + self.QUESTION[0] for n in num_list],
                                             [n + self.ANSWER[0] for n in num_list],
                                             [n + self.RESOLVE[0] for n in num_list])
            short_list = self.run_mode(mode, [n + self.QUESTION[1] for n in num_list],
                                             [n + self.ANSWER[1] for n in num_list],
                                             [n + self.RESOLVE[1] for n in num_list])

            self.doc_info[row['试卷名称']] = {
                'keyword': keyword,
                'mode': mode,
                'volume': volume,
                'num_list': num_list,
                'final_list': final_list,
                'short_list': short_list,
            }

    def write_to_excel(self):
        # Workbook 1
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '目录'
        for i, col in enumerate(self.content_cols):
            ws.cell(1, i + 1).value = col
        temp_row = 2
        for k, v in self.doc_info.items():
            ws.cell(temp_row, 1).value = k
            temp_row += 1
            for i, n in enumerate(v['num_list']):
                for _ in range(5, 7):
                    ws.cell(temp_row, _).value = n
                for _ in range(7, 10):
                    ws.cell(temp_row, _).value = i + 1
                temp_row += 1
        self.format_worksheet(ws)
        wb.save("目录.xlsx")
        # Workbook 2
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '重命名表格'
        temp_row = 1
        for k, v in self.doc_info.items():
            ws.cell(temp_row, 3).value = k
            for i in range(len(v['final_list'])):
                ws.cell(temp_row, 1).value = v['final_list'][i]
                ws.cell(temp_row, 2).value = v['short_list'][i]
                temp_row += 1
        self.format_worksheet(ws)
        wb.save("重命名表格.xlsx")
        del wb, ws
        gc.collect()

    def format_worksheet(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    def run_mode(self, mode, ql, al, rl):
        """ direct run with specific mode """
        return self.find_func(mode, ql, al, rl)

    def run_input(self):
        tips = """模式：
    ①题目1-答案1-解析1-题目2-答案2-解析2……
    ②题目1-解析1-答案1-题目2-解析2-答案2……
    ③题目1-题目2……答案1-解析1-答案2-解析2……
    ④题目1-题目2……解析1-答案1-解析2-答案2-……
    ⑤题目1-题目2……答案1-答案2……解析1-解析2……
    ⑥题目1-题目2-题目3……
    ⑦答案1-答案2-答案3……
    ⑧解析1-解析2-解析3……
    ⑨题目1-答案1-题目2-答案2……
    ⑩题目1-题目2……答案1-答案2……"""
        while True:
            print(tips)
            n = input("请输入模式编号(整数)：")
            try:
                n = int(n)
            except:
                print("请输入整数")
                continue

            result = self.find_func(n, self.ql, self.al, self.rl)
            if not isinstance(result, list):
                print("模式不存在，请重新输入")
                continue

            for i in result:
                print(i)

            with open("mode_{}_output.txt".format(n), "w") as f:
                for i in result:
                    f.write(i + '\n')

def main():
    rn = RenameNumber()
    # r = rn.run_mode(9)
    # print(r)
    try:
        rn.get_document_info()
        rn.write_to_excel()
    except Exception as e:
        print(str(e))
    finally:
        input("\n请按任何键以继续...")

main()